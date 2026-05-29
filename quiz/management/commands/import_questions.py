from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from quiz.models import Question, QuizCategory


class Command(BaseCommand):
    help = "Import quiz questions from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]

        workbook = load_workbook(file_path)
        sheet = workbook.active

        imported_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                category_name,
                category_slug,
                section,
                title,
                option_1,
                option_2,
                option_3,
                option_4,
                correct_answer,
                points,
                explanation,
                is_published,
            ) = row

            if not title:
                continue

            category, created = QuizCategory.objects.get_or_create(
                slug=category_slug,
                defaults={
                    "name": category_name,
                },
            )

            Question.objects.create(
                category=category,
                section=section,
                title=title,
                option_1=option_1,
                option_2=option_2,
                option_3=option_3,
                option_4=option_4 or "",
                correct_answer=int(correct_answer),
                points=int(points or 3),
                explanation=explanation or "",
                is_published=bool(is_published),
            )

            imported_count += 1

        self.stdout.write(
            self.style.SUCCESS(f"{imported_count} questions imported successfully.")
        )